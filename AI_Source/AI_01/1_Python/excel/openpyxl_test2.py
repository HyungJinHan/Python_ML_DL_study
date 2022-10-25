from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("Diary", 0) # 0이면 첫 번째 위치에 삽입, 생략 시 마지막 위치에 삽입

data = [('홍길동', 80, 90, 70), ('이기자', 90, 60, 80), ('강기자', 80, 80, 70)]

r = 0 # openpyxl는 첫 번째 행이 0부터 시작
c = 0 # openpyxl는 첫 번째 열이 0부터 시작

for name, kor, eng, math in data:
    ws.cell(row=r, column=c).value = name
    # r(1)행 c(1)열 위치의 데이터를 name으로 언패킹
    ws.cell(row=r, column=c + 1).value = kor
    ws.cell(row=r, column=c + 2).value = eng
    ws.cell(row=r, column=c + 3).value = math
    r += 1

""" 다른 방법 1
for name, kor, eng, math in data:
    ws['A' + str(r)].value = name
    ws['B' + str(r)].value = kor
    ws['C' + str(r)].value = eng
    ws['D' + str(r)].value = math
    r += 1
"""

""" 다른 방법 2
columnChar = 'A'
for name, kor, eng, math in data:
    ws[columnChar + str(r)].value = name
    # ws[columnChar + str(r)].value = name.encode(encoding='utf_8', errors='ignore')
    ws[chr(ord(columnChar) + 1) + str(r)].value = kor
    ws[chr(ord(columnChar) + 2) + str(r)].value = eng
    ws[chr(ord(columnChar) + 3) + str(r)].value = math
    r += 1
"""

""" 다른 방법 3
columnChar = 'A'
for val in data:
    for i in range(0, 4):
        ws[chr(ord(columnChar) + i) + str(r)].value = val[i]
    r += 1
"""

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save('openpyxl2.xlsx')
wb.close()