from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

ws.title = 'Test'
ws['B2'] = 42
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([4, 4, 4])
wb.save('openpyxl1.xlsx')
wb.close()

wb = load_workbook(filename='openpyxl1.xlsx')
ws = wb.active
ws['A1'] = 42
ws.cell(row=1, column=3).value = 333
ws.append(['aaa', 'bbb', 'ccc']) # 새로운 행 추가

print(ws['A1'].value)
print(ws['A2'].value) # 저장 값이 없다면 None

ws2 = wb['Test'] # Test 시트 선택
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save('openpyxl1.xlsx')