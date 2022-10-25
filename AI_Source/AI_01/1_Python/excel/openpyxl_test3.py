from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("Chart", 0) # 0이면 첫 번째 위치에 삽입, 생략 시 마지막 위치에 삽입

ws.merge_cells('A1:D1') # 셀 병합
# ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
ws['A1'] = '성적표'
ws['A1'].font = Font(name='맑은 고딕', size=15, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
# ws.alignment = Alignment(horizontal='center', vertical='center')

ws.append(['이름', '국어', '영어', '수학'])
ws.append(['한형진', 100, 100, 100])
ws.append(['전우진', 90, 90, 90])
ws.append(['김성환', 80, 80, 80])
ws.append(['김건', 70, 70, 70])
ws.append(['문경현', 60, 60, 60])
ws.append(['황기성', 50, 50, 50])

wb.save('openpyxl_chart.xlsx')

barchart = BarChart()
barchart.title = '성적표'
barchart.x_axis.title = '이름'
barchart.y_axis.title = '점수'

data = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=8)
# 차트 만들때 사용될 데이터 범위 지정
cate = Reference(ws, min_col=1, min_row=3, max_row=8)
# 데이터 카데고리 범위 지정

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(cate)
barchart.shape = 1

ws.add_chart(barchart, 'F1')
# F1은 차트가 만들어질 위치
wb.save('openpyxl_chart.xlsx')

wb.close()